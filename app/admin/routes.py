from flask import render_template, redirect, url_for, flash, request, current_app
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from app.admin import bp
from app import db
from app.models import User, Admin, Event, Meeting, QRCode, Attendance, Region, Team, AccessRequest
from flask import send_file, make_response, abort
import io
import qrcode
import secrets
from datetime import datetime


@bp.route('/login', methods=['GET', 'POST'])
def login():
    from flask import current_app
    if request.method == 'POST':
        telefone = request.form.get('telefone', '')
        # normalize phone digits
        telefone = ''.join(ch for ch in telefone if ch.isdigit())
        current_app.logger.debug(f"admin login attempt with telefone={telefone}")
        password = request.form.get('password')
        try:
            user = User.query.filter_by(telefone=telefone).first()
            # if the phone isn't found, try to locate by normalizing stored values
            if user is None:
                for u in User.query.all():
                    norm = ''.join(ch for ch in u.telefone if ch.isdigit())
                    if norm == telefone:
                        user = u
                        # persist normalized value for future logins
                        u.telefone = telefone
                        db.session.commit()
                        current_app.logger.debug(f"normalized stored telefone for user {u.id}")
                        break
        except Exception as e:
            # could be operational error (DB not reachable)
            flash('Erro de conexão ao banco de dados, verifique se o serviço está ativo.', 'danger')
            return render_template('login.html')
        # add debug info for failure cases
        if user is None:
            current_app.logger.debug("login failed: no matching user")
        elif not user.admin_record:
            current_app.logger.debug(f"login failed: user {user.telefone} exists but is not admin")
        else:
            admin = user.admin_record
            if admin.check_password(password):
                login_user(user)
                return redirect(url_for('admin.dashboard'))
            else:
                current_app.logger.debug("login failed: password mismatch for admin user")
        flash('Usuário ou senha inválidos', 'danger')
    return render_template('login.html')


@bp.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('main.index'))


@bp.route('/')
@login_required
def dashboard():
    # filtering parameters
    participant_q = request.args.get('participant', '').strip().lower()
    region_id = request.args.get('region_id')
    regions = Region.query.order_by(Region.nome).all()
    # open QR codes (active and event not finished)
    today = datetime.utcnow().date()
    open_qrs = [q for q in QRCode.query.filter_by(active=True).all() if q.meeting.event.data_final >= today]
    # meeting statistics
    stats = []
    chart_labels = []
    chart_totals = []
    # region totals for pie chart
    region_counts = {}
    # iterate by event -> meetings sorted by date
    events = Event.query.order_by(Event.data_inicial).all()
    # region-by-meeting counts will support another chart
    region_meeting = {}  # region -> list of totals per meeting (aligned with chart_labels)
    for ev in events:
        meetings = sorted(ev.meetings, key=lambda m: m.data)
        for idx, mt in enumerate(meetings):
            # apply filters: participant and region
            if participant_q:
                matched = any(participant_q in a.user.nome.lower() or participant_q in a.user.telefone for a in mt.attendances)
                if not matched:
                    continue
            if region_id:
                matched = any(str(a.user.region_id) == str(region_id) for a in mt.attendances)
                if not matched:
                    continue
            total = len(mt.attendances)
            # accumulate bar chart
            chart_labels.append(f"{ev.nome} - {mt.titulo or mt.data.strftime('%d/%m/%Y')}")
            chart_totals.append(total)
            by_region = {}
            for att in mt.attendances:
                cor = att.user.cor
                by_region[cor] = by_region.get(cor, 0) + 1
                region_counts[cor] = region_counts.get(cor, 0) + 1
            # record per-region data for this meeting
            for cor, cnt in by_region.items():
                region_meeting.setdefault(cor, []).append(cnt)
            # ensure zeros for regions with no attendees this meeting
            for cor in list(region_counts.keys()):
                if cor not in by_region:
                    region_meeting.setdefault(cor, []).append(0)
            new = missing = 0
            prev = meetings[idx - 1] if idx > 0 else None
            if prev:
                prev_users = {a.user_id for a in prev.attendances}
                curr_users = {a.user_id for a in mt.attendances}
                new = len(curr_users - prev_users)
                missing = len(prev_users - curr_users)
            stats.append({'event': ev, 'meeting': mt, 'total': total,
                          'by_region': by_region, 'new': new, 'missing': missing})
    return render_template('dashboard.html', open_qrs=open_qrs, stats=stats,
                           chart_labels=chart_labels, chart_totals=chart_totals,
                           region_labels=list(region_counts.keys()),
                           region_totals=list(region_counts.values()),
                           region_meeting=region_meeting,
                           participant_q=participant_q, regions=regions, region_id=region_id)


@bp.route('/dashboard/export/<fmt>')
@login_required

def export_dashboard(fmt):
    # replicate logic from dashboard for stats without filters
    stats = []
    region_counts = {}
    events = Event.query.order_by(Event.data_inicial).all()
    for ev in events:
        meetings = sorted(ev.meetings, key=lambda m: m.data)
        for idx, mt in enumerate(meetings):
            total = len(mt.attendances)
            by_region = {}
            for att in mt.attendances:
                cor = att.user.cor
                by_region[cor] = by_region.get(cor, 0) + 1
                region_counts[cor] = region_counts.get(cor, 0) + 1
            new = missing = 0
            prev = meetings[idx - 1] if idx > 0 else None
            if prev:
                prev_users = {a.user_id for a in prev.attendances}
                curr_users = {a.user_id for a in mt.attendances}
                new = len(curr_users - prev_users)
                missing = len(prev_users - curr_users)
            stats.append({'event': ev, 'meeting': mt, 'total': total,
                          'by_region': by_region, 'new': new, 'missing': missing})
    # prepare rows for export
    rows = []
    for s in stats:
        row = {
            'Evento': s['event'].nome,
            'Reunião': s['meeting'].titulo or s['meeting'].data.strftime('%d/%m/%Y'),
            'Total': s['total'],
            'Novos': s['new'],
            'Faltaram': s['missing'],
        }
        # flatten region breakdown
        for cor, cnt in s['by_region'].items():
            row[f'Região {cor}'] = cnt
        rows.append(row)
    if fmt == 'xlsx':
        import pandas as pd, io
        df = pd.DataFrame(rows)
        bio = io.BytesIO()
        df.to_excel(bio, index=False)
        bio.seek(0)
        return send_file(bio, download_name="dashboard_stats.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        import io
        bio = io.BytesIO()
        c = canvas.Canvas(bio, pagesize=letter)
        text = c.beginText(40, 750)
        text.textLine("Dashboard statistics")
        text.textLine('')
        for s in stats:
            text.textLine(f"{s['event'].nome} - {s['meeting'].titulo or s['meeting'].data.strftime('%d/%m/%Y')}: {s['total']} presenças")
        c.drawText(text)
        c.showPage()
        c.save()
        bio.seek(0)
        return send_file(bio, download_name="dashboard_stats.pdf",
                         as_attachment=True,
                         mimetype='application/pdf')
    else:
        abort(404)
    for ev in events:
        meetings = sorted(ev.meetings, key=lambda m: m.data)
        for idx, mt in enumerate(meetings):
            total = len(mt.attendances)
            by_region = {}
            for att in mt.attendances:
                cor = att.user.cor
                by_region[cor] = by_region.get(cor, 0) + 1
            new = missing = 0
            prev = meetings[idx - 1] if idx > 0 else None
            if prev:
                prev_users = {a.user_id for a in prev.attendances}
                curr_users = {a.user_id for a in mt.attendances}
                new = len(curr_users - prev_users)
                missing = len(prev_users - curr_users)
            stats.append({'event': ev, 'meeting': mt, 'total': total,
                          'by_region': by_region, 'new': new, 'missing': missing})
    return render_template('dashboard.html', open_qrs=open_qrs, stats=stats)


# ---------- utility actions ----------

@bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    # configure globals such as API token (other integrations removed)
    from app.models import Setting
    if request.method == 'POST':
        for key in ('API_TOKEN',):
            val = request.form.get(key.lower())
            if val is not None:
                Setting.set(key, val.strip())
        flash('Configurações atualizadas', 'success')
        return redirect(url_for('admin.settings'))
    # prepare existing values
    vals = {k.lower(): Setting.get(k,'') for k in ('API_TOKEN',)}
    return render_template('settings.html', **vals)




@bp.route('/whatsapp', methods=['GET', 'POST'])
@login_required
def whatsapp():
    # form that builds a direct WhatsApp web link; no external service required
    link = None
    # allow prefilling via GET params
    to = ''
    body = ''
    if request.method == 'POST':
        to = request.form.get('to', '').strip()
        body = request.form.get('body', '').strip()
    else:
        to = request.args.get('to', '').strip()
        body = request.args.get('body', '').strip()
    if request.method == 'POST' or (to and body):
        if to and body:
            from urllib.parse import quote_plus
            # remove non-digits and ensure international format without '+'
            clean = ''.join(ch for ch in to if ch.isdigit())
            link = f"https://wa.me/{clean}?text={quote_plus(body)}"
        else:
            flash('Todos os campos são obrigatórios', 'danger')
    return render_template('whatsapp.html', link=link, to=to, body=body)

@bp.route('/clear_data', methods=['POST'])
@login_required
def clear_data():
    # only original admin allowed
    if not current_user.admin_record.is_original:
        flash('Permissão negada', 'danger')
        return redirect(url_for('admin.dashboard'))
    from sqlalchemy import text
    # delete all except original admin user
    db.session.execute(text('DELETE FROM attendance'))
    db.session.execute(text('DELETE FROM qr_code'))
    db.session.execute(text('DELETE FROM meeting'))
    db.session.execute(text('DELETE FROM event'))
    db.session.execute(text('DELETE FROM team_member'))
    db.session.execute(text('DELETE FROM team'))
    # do not remove users with admin original
    db.session.execute(text("DELETE FROM admin WHERE is_original = 0"))
    db.session.execute(text("DELETE FROM user WHERE id NOT IN (SELECT user_id FROM admin WHERE is_original=1)"))
    db.session.commit()
    flash('Dados do sistema removidos', 'success')
    return redirect(url_for('admin.dashboard'))

# ---------- event & meeting management ----------

@bp.route('/events')
@login_required
def list_events():
    page = request.args.get('page', 1, type=int)
    per_page = 12
    pagination = Event.query.order_by(Event.data_inicial.desc()).paginate(page=page, per_page=per_page, error_out=False)
    events = pagination.items
    return render_template('events.html', events=events, pagination=pagination)


@bp.route('/events/<int:event_id>/delete', methods=['GET','POST'])
@login_required
def delete_event(event_id):
    ev = Event.query.get_or_404(event_id)
    if request.method == 'GET':
        flash('Use o botão de exclusão na lista para remover o evento.', 'info')
        return redirect(url_for('admin.list_events'))
    # Removing the Event object will cascade to its meetings and their teams
    # thanks to SQLAlchemy relationships defined with cascade='all, delete-orphan'.
    db.session.delete(ev)
    db.session.commit()
    flash('Evento excluído', 'success')
    return redirect(url_for('admin.list_events'))


@bp.route('/events/new', methods=['GET', 'POST'])
@login_required
def new_event():
    if request.method == 'POST':
        nome = request.form.get('nome')
        data = request.form.get('data')
        dataf = request.form.get('data_final')
        if not nome or not data or not dataf:
            flash('Título, data inicial e data final são obrigatórios', 'danger')
            return redirect(url_for('admin.new_event'))
        d1 = datetime.fromisoformat(data).date()
        d2 = datetime.fromisoformat(dataf).date()
        if d2 < d1:
            flash('Data final não pode ser anterior à inicial', 'danger')
            return redirect(url_for('admin.new_event'))
        ev = Event(nome=nome.strip().upper(), data_inicial=d1, data_final=d2)
        db.session.add(ev)
        db.session.commit()
        return redirect(url_for('admin.list_events'))
    return render_template('new_event.html')

@bp.route('/events/<int:event_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_event(event_id):
    ev = Event.query.get_or_404(event_id)
    if request.method == 'POST':
        nome = request.form.get('nome')
        if nome:
            ev.nome = nome.strip().upper()
        data = request.form.get('data')
        dataf = request.form.get('data_final')
        if data:
            ev.data_inicial = datetime.fromisoformat(data).date()
        if dataf:
            ev.data_final = datetime.fromisoformat(dataf).date()
        if ev.data_final < ev.data_inicial:
            flash('Data final não pode ser anterior à inicial', 'danger')
            return redirect(url_for('admin.edit_event', event_id=event_id))
        db.session.commit()
        flash('Evento atualizado', 'success')
        return redirect(url_for('admin.list_events'))
    return render_template('edit_event.html', event=ev)



@bp.route('/events/<int:event_id>')
@login_required
def event_detail(event_id):
    ev = Event.query.get_or_404(event_id)
    return render_template('event_detail.html', event=ev, datetime=datetime)
# region management
@bp.route('/regions')
@login_required
def list_regions():
    regions = Region.query.order_by(Region.nome).all()
    return render_template('regions.html', regions=regions)

@bp.route('/access-requests')
@login_required
def list_access_requests():
    reqs = AccessRequest.query.order_by(AccessRequest.criado_em.desc()).all()
    return render_template('access_requests.html', requests=reqs)

@bp.route('/qrcodes')
@login_required
def list_qrcodes():
    # group by event->meeting->qrcode
    events = Event.query.order_by(Event.data_inicial).all()
    return render_template('qrcodes.html', events=events)

@bp.route('/access-requests/<int:req_id>/approve', methods=['POST'])
@login_required
def approve_access_request(req_id):
    r = AccessRequest.query.get_or_404(req_id)
    if not r.approved:
        user = User(telefone=r.telefone, nome=r.nome, cor=r.region.nome if r.region else '', region=r.region)
        db.session.add(user)
        db.session.commit()
        # create admin credentials so can log in (password already hashed)
        admin = Admin(user=user, is_original=False, password_hash=r.password_hash)
        db.session.add(admin)
        r.approved = True
        db.session.commit()
        flash('Solicitação aprovada, usuário e administrador criados', 'success')
    else:
        flash('Solicitação já aprovada', 'info')
    return redirect(url_for('admin.list_access_requests'))

@bp.route('/regions/new', methods=['GET', 'POST'])
@login_required
def new_region():
    if request.method == 'POST':
        name = request.form.get('nome')
        if not name:
            flash('Nome é obrigatório', 'danger')
            return redirect(url_for('admin.new_region'))
        name = name.strip().upper()
        if Region.query.filter_by(nome=name).first():
            flash('Região já existe', 'danger')
            return redirect(url_for('admin.new_region'))
        r = Region(nome=name)
        db.session.add(r)
        db.session.commit()
        flash('Região criada', 'success')
        return redirect(url_for('admin.list_regions'))
    return render_template('new_region.html')

@bp.route('/regions/<int:region_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_region(region_id):
    r = Region.query.get_or_404(region_id)
    if request.method == 'POST':
        name = request.form.get('nome', '')
        r.nome = name.strip().upper()
        db.session.commit()
        flash('Região atualizada', 'success')
        return redirect(url_for('admin.list_regions'))
    return render_template('edit_region.html', region=r)

@bp.route('/regions/<int:region_id>/delete', methods=['POST'])
@login_required
def delete_region(region_id):
    r = Region.query.get_or_404(region_id)
    # nullify users
    for u in r.users:
        u.region_id = None
    db.session.delete(r)
    db.session.commit()
    flash('Região excluída', 'success')
    return redirect(url_for('admin.list_regions'))
# user management

@bp.route('/users')
@login_required
def list_users():
    # filtering
    name_q = request.args.get('name', '').strip()
    region_id = request.args.get('region_id')
    page = request.args.get('page', 1, type=int)
    per_page = 15
    # pre-migrate any users whose region is stored in cor but region_id is null
    # (ensures filter will catch them)
    for u in User.query.filter(User.region_id.is_(None), User.cor.isnot(None)).all():
        r = Region.query.filter_by(nome=u.cor).first()
        if r:
            u.region_id = r.id
    db.session.commit()
    # determine whether original admin should be visible
    show_original = False
    if current_user.is_authenticated:
        show_original = (current_user.telefone == '14981364342' or
                         (current_user.admin_record and current_user.admin_record.is_original))
    # build query
    query = User.query
    if not show_original:
        # exclude original administrator from everyone else
        query = query.filter(~User.admin_record.has(is_original=True))
    if name_q:
        query = query.filter(User.nome.ilike(f"%{name_q}%"))
    if region_id:
        # allow matches either by region_id or by cor (in case migration not applied)
        query = query.filter(
            (User.region_id == region_id) | (User.cor.ilike(
                Region.query.with_entities(Region.nome).filter_by(id=region_id).scalar_subquery()
            ))
        )
    pagination = query.order_by(User.nome).paginate(page=page, per_page=per_page, error_out=False)
    users = pagination.items
    regions = Region.query.order_by(Region.nome).all()
    return render_template('users.html', users=users, regions=regions,
                           name_q=name_q, region_id=region_id, pagination=pagination)

@bp.route('/users/new', methods=['GET', 'POST'])
@login_required
def new_user():
    regions = Region.query.order_by(Region.nome).all()
    if request.method == 'POST':
        telefone = request.form.get('telefone','').strip()
        # normalize to digits only to match login behavior
        telefone = ''.join(ch for ch in telefone if ch.isdigit())
        nome = request.form.get('nome','').strip()
        region_id = request.form.get('region_id') or None
        cor = request.form.get('cor','').strip() or ''
        if not telefone or not nome:
            flash('Telefone e nome são obrigatórios', 'danger')
            return redirect(url_for('admin.new_user'))
        if User.query.filter_by(telefone=telefone).first():
            flash('Telefone já cadastrado', 'danger')
            return redirect(url_for('admin.new_user'))
        u = User(telefone=telefone, nome=nome, cor=cor or '-', region_id=region_id)
        db.session.add(u)
        db.session.commit()
        flash('Usuário criado', 'success')
        return redirect(url_for('admin.list_users'))
    return render_template('new_user.html', regions=regions)

@bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    # if region not set but cor matches, assign
    if not user.region_id and user.cor:
        r = Region.query.filter_by(nome=user.cor).first()
        if r:
            user.region_id = r.id
    regions = Region.query.order_by(Region.nome).all()
    if request.method == 'POST':
        user.nome = request.form.get('nome')
        user.cor = request.form.get('cor')
        user.region_id = request.form.get('region_id') or None
        db.session.commit()
        flash('Usuário atualizado', 'success')
        return redirect(url_for('admin.list_users'))
    return render_template('edit_user.html', user=user, regions=regions)

@bp.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    # prevent removal of original admin contact
    if user.admin_record and user.admin_record.is_original:
        flash('Não é permitido excluir o administrador original', 'danger')
        return redirect(url_for('admin.list_users'))
    # perform raw SQL deletes to fully remove user and related records
    from sqlalchemy import text
    # order matters due to FK constraints
    db.session.execute(text('DELETE FROM attendance WHERE user_id = :uid'), {'uid': user.id})
    db.session.execute(text('DELETE FROM team_member WHERE user_id = :uid'), {'uid': user.id})
    db.session.execute(text('DELETE FROM admin WHERE user_id = :uid'), {'uid': user.id})
    db.session.execute(text('DELETE FROM user WHERE id = :uid'), {'uid': user.id})
    db.session.commit()
    flash('Usuário excluído', 'success')
    return redirect(url_for('admin.list_users'))


@bp.route('/events/<int:event_id>/meetings/new', methods=['GET', 'POST'])
@login_required
def new_meeting(event_id):
    ev = Event.query.get_or_404(event_id)
    now = datetime.utcnow().date()
    if now > ev.data_final:
        flash('Evento encerrado, não é possível adicionar reuniões', 'danger')
        return redirect(url_for('admin.event_detail', event_id=event_id))
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        data = request.form.get('data')
        if not titulo or not data:
            flash('Título e data são obrigatórios', 'danger')
            return redirect(url_for('admin.new_meeting', event_id=event_id))
        d = datetime.fromisoformat(data).date()
        if d < ev.data_inicial or d > ev.data_final:
            flash('Data da reunião fora do intervalo do evento', 'danger')
            return redirect(url_for('admin.new_meeting', event_id=event_id))
        special = bool(request.form.get('special'))
        mt = Meeting(event=ev, titulo=titulo.strip().upper(), data=d, special=special)
        db.session.add(mt)
        db.session.commit()
        return redirect(url_for('admin.event_detail', event_id=event_id))
    return render_template('new_meeting.html', event=ev)

@bp.route('/meetings/<int:meeting_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_meeting(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    ev = mt.event
    if request.method == 'POST':
        titulo = request.form.get('titulo')
        if titulo:
            mt.titulo = titulo.strip().upper()
        data = request.form.get('data')
        if data:
            new_date = datetime.fromisoformat(data).date()
            if new_date < ev.data_inicial or new_date > ev.data_final:
                flash('Data da reunião fora do intervalo do evento', 'danger')
                return redirect(url_for('admin.edit_meeting', meeting_id=meeting_id))
            mt.data = new_date
        special = bool(request.form.get('special'))
        mt.special = special
        db.session.commit()
        flash('Reunião atualizada', 'success')
        return redirect(url_for('admin.event_detail', event_id=mt.event_id))
    return render_template('edit_meeting.html', meeting=mt)


@bp.route('/meetings/<int:meeting_id>')
@login_required
def meeting_detail(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    return render_template('meeting_detail.html', meeting=mt)


# compute next meeting date and type for event pattern
from datetime import timedelta

def _next_meeting(mt):
    ev = mt.event
    start = mt.data + timedelta(days=1)
    # find next Tuesday
    cur = start
    while cur <= ev.data_final:
        if cur.weekday() == 1:  # Tuesday
            return cur, 'preparação'
        cur += timedelta(days=1)
    # no Tuesday found, use Thursday before end
    final = ev.data_final - timedelta(days=(ev.data_final.weekday() - 3) % 7)
    if final > mt.data:
        return final, 'unção'
    return None, None


@bp.route('/meetings/<int:meeting_id>/missed')
@login_required
def meeting_missed(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    # users who attended any meeting of this event
    ev = mt.event
    user_ids = {a.user_id for m in ev.meetings for a in m.attendances}
    present_ids = {a.user_id for a in mt.attendances}
    missed_ids = user_ids - present_ids
    missed = User.query.filter(User.id.in_(missed_ids)).order_by(User.nome).all()
    next_date, next_type = _next_meeting(mt)
    return render_template('meeting_missed.html', meeting=mt, missed=missed,
                           next_date=next_date, next_type=next_type)

# team management for special meetings
@bp.route('/meetings/<int:meeting_id>/teams', methods=['GET','POST'])
@login_required
def meeting_teams(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    if not mt.special:
        flash('Somente reuniões especiais aceitam equipes', 'warning')
        return redirect(url_for('admin.meeting_detail', meeting_id=meeting_id))
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if name:
            if not Team.query.filter_by(meeting_id=mt.id, nome=name).first():
                t = Team(meeting=mt, nome=name)
                db.session.add(t)
                db.session.commit()
                flash('Equipe criada', 'success')
            else:
                flash('Equipe já existe', 'info')
        return redirect(url_for('admin.meeting_teams', meeting_id=meeting_id))
    teams = Team.query.filter_by(meeting=mt).all()
    return render_template('meeting_teams.html', meeting=mt, teams=teams)

@bp.route('/meetings/<int:meeting_id>/teams/<int:team_id>/delete', methods=['POST'])
@login_required
def delete_team(meeting_id, team_id):
    t = Team.query.get_or_404(team_id)
    if t.meeting_id != meeting_id:
        abort(404)
    db.session.delete(t)
    db.session.commit()
    flash('Equipe removida', 'success')
    return redirect(url_for('admin.meeting_teams', meeting_id=meeting_id))



@bp.route('/teams/<int:team_id>/members', methods=['GET','POST'])
@login_required
def manage_team_members(team_id):
    team = Team.query.get_or_404(team_id)
    mt = team.meeting
    # get users who attended this meeting
    # hide original admin from member selection unless he is the current user
    user_ids = [a.user_id for a in mt.attendances if not (a.user.admin_record and a.user.admin_record.is_original and (not (current_user.is_authenticated and current_user.telefone=='14981364342')))]
    # compute attendance counts across all meetings for these users
    from sqlalchemy import func
    counts = (db.session.query(User.id, func.count(Attendance.id).label('cnt'))
              .join(Attendance).filter(User.id.in_(user_ids))
              .group_by(User.id).subquery())
    users = (db.session.query(User, counts.c.cnt)
             .join(counts, User.id == counts.c.id)
             .order_by(counts.c.cnt.desc(), User.nome).all())
    if request.method == 'POST':
        selected = request.form.getlist('user_id')
        team.users = User.query.filter(User.id.in_(selected)).all()
        # leader selection (may be empty or not in users)
        leader_val = request.form.get('leader_id')
        if leader_val:
            leader = User.query.get(int(leader_val))
            if leader in team.users:
                team.leader = leader
        else:
            team.leader = None
        db.session.commit()
        flash('Membros atualizados', 'success')
        return redirect(url_for('admin.manage_team_members', team_id=team_id))
    return render_template('meeting_team_members.html', team=team, users=users)

@bp.route('/meetings/<int:meeting_id>/attendance')
@login_required
def attendance_list(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    return render_template('attendance.html', meeting=mt)

@bp.route('/attendance/<int:att_id>/delete', methods=['POST'])
@login_required
def delete_attendance(att_id):
    att = Attendance.query.get_or_404(att_id)
    mt_id = att.meeting_id
    db.session.delete(att)
    db.session.commit()
    flash('Presença removida', 'success')
    return redirect(url_for('admin.attendance_list', meeting_id=mt_id))

@bp.route('/meetings/<int:meeting_id>/attendance/export/<fmt>')
@login_required
def export_attendance(meeting_id, fmt):
    mt = Meeting.query.get_or_404(meeting_id)
    # prepare data rows – sort by user name
    attendances = sorted(mt.attendances, key=lambda a: a.user.nome)
    rows = []
    for idx, a in enumerate(attendances, start=1):
        rows.append({
            'Código': idx,
            'Nome': a.user.nome,
            'Telefone': a.user.telefone,
            'Região': a.user.cor,
            'Confirmado em': a.confirmado_em.strftime('%d/%m/%Y %H:%M')
        })
    # create a base filename from meeting title/date
    clean_name = (mt.titulo or mt.data.strftime('%d-%m-%Y')).replace(' ', '_')
    if fmt == 'xlsx':
        import pandas as pd, io
        df = pd.DataFrame(rows)
        bio = io.BytesIO()
        df.to_excel(bio, index=False)
        bio.seek(0)
        return send_file(bio, download_name=f"{clean_name}.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        # build a table with headers and totals
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        bio = io.BytesIO()
        doc = SimpleDocTemplate(bio, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        # include logo for PDF
        try:
            import os
            from reportlab.platypus import Image
            logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
            logo = Image(logo_path, width=100, height=100)
            elements.append(logo)
            elements.append(Spacer(1,6))
        except Exception:
            pass
        # heading structure: main title, then meeting title and date
        elements.append(Paragraph("Lista de presença", styles['Heading1']))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(mt.titulo or '', styles['Heading3']))
        elements.append(Paragraph(mt.data.strftime('%d/%m/%Y'), styles['Heading4']))
        elements.append(Spacer(1, 12))
        # build table data
        data = [['Código', 'Nome', 'Telefone', 'Região', 'Confirmado em']]
        for row in rows:
            data.append([row['Código'], row['Nome'], row['Telefone'], row['Região'], row['Confirmado em']])
        # totals row inside table
        data.append(['', '', '', 'Total presentes', len(rows)])
        t = Table(data, repeatRows=1)
        tbl_style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black),
                                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                                ('SPAN', (-2, -1), (-1, -1))])
        t.setStyle(tbl_style)
        elements.append(t)
        elements.append(Spacer(1,12))
        # summary text below table
        elements.append(Paragraph(f"Total de presentes: {len(rows)}", styles['Normal']))
        def footer(canvas, doc):
            width, height = letter
            canvas.setFont("Helvetica", 8)
            canvas.drawString(40, 20, "Primeira Igreja Batista de Marília")
            canvas.drawCentredString(width/2, 20, "Face a Face de Homens")
            canvas.drawRightString(width-40, 20, str(doc.page))
        doc.build(elements, onFirstPage=footer, onLaterPages=footer)
        bio.seek(0)
        return send_file(bio, download_name=f"{clean_name}.pdf",
                         as_attachment=True,
                         mimetype='application/pdf')
    else:
        abort(404)


# export across all meetings of an event with status matrix
@bp.route('/events/<int:event_id>/attendance/export/<fmt>')
@login_required
def export_event_attendance(event_id, fmt):
    ev = Event.query.get_or_404(event_id)
    meetings = sorted(ev.meetings, key=lambda m: m.data)
    # build set of users who participated in any meeting
    user_ids = set(a.user_id for m in meetings for a in m.attendances)
    users = User.query.filter(User.id.in_(user_ids)).order_by(User.nome).all()
    # header columns: Código, Nome, Telefone, Região, then one column per meeting date
    date_cols = [m.data.strftime('%d/%m/%Y') for m in meetings]
    rows = []
    for idx, u in enumerate(users, start=1):
        row = {
            'Código': idx,
            'Nome': u.nome,
            'Telefone': u.telefone,
            'Região': u.cor
        }
        for m in meetings:
            attended = any(a.user_id == u.id for a in m.attendances)
            row[m.data.strftime('%d/%m/%Y')] = 'PRESENTE' if attended else 'FALTOU'
        rows.append(row)
    # filenames based on event name
    safe_event = ev.nome.replace(' ', '_')
    if fmt == 'xlsx':
        import pandas as pd, io
        from openpyxl.styles import PatternFill
        df = pd.DataFrame(rows)
        # append totals row if any
        if not df.empty:
            total_values = {}
            for col in date_cols:
                present = (df[col] == 'PRESENTE').sum()
                absent = (df[col] == 'FALTOU').sum()
                total_values[col] = f"P {present} / F {absent}"
            total_values.update({'Código':'','Nome':'','Telefone':'','Região':'Totais:'})
            df = pd.concat([df, pd.DataFrame([total_values])], ignore_index=True)
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Resumo')
            wb = writer.book
            ws = writer.sheets['Resumo']
            # color cells based on status
            green = PatternFill(start_color='00C6EFCE', fill_type='solid')
            red = PatternFill(start_color='00FFC7CE', fill_type='solid')
            # find status columns indexes
            for col_idx, col in enumerate(df.columns, start=1):
                if col in date_cols:
                    for row_idx in range(2, len(df) + 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value == 'PRESENTE':
                            cell.fill = green
                        elif cell.value == 'FALTOU':
                            cell.fill = red
        bio.seek(0)
        return send_file(bio, download_name=f"{safe_event}.xlsx",
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif fmt == 'pdf':
        # create a table with colored cells using reportlab
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        import io
        bio = io.BytesIO()
        doc = SimpleDocTemplate(bio, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        # include logo
        try:
            import os
            logo_path = os.path.join(current_app.root_path, 'static', 'logo.png')
            logo = Image(logo_path, width=100, height=100)
            elements.append(logo)
            elements.append(Spacer(1,6))
        except Exception:
            pass
        # heading
        elements.append(Paragraph('Lista de presença', styles['Heading1']))
        elements.append(Spacer(1,6))
        elements.append(Paragraph(ev.nome.upper(), styles['Heading3']))
        elements.append(Spacer(1,12))
        data = []
        header = ['Código', 'Nome', 'Telefone', 'Região'] + date_cols
        data.append(header)
        for row in rows:
            data.append([row[h] for h in header])
        # summary totals row
        if rows:
            totals = ['','', '', 'Totais:']
            for col in date_cols:
                present = sum(1 for r in rows if r[col] == 'PRESENTE')
                absent = sum(1 for r in rows if r[col] == 'FALTOU')
                totals.append(f"P {present} / F {absent}")
            data.append(totals)
        t = Table(data, repeatRows=1)
        tbl_style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black)])
        # apply backgrounds to status cells
        for r, row in enumerate(data[1:], start=1):
            for c, col in enumerate(header):
                if col in date_cols:
                    val = row[c]
                    if val == 'PRESENTE':
                        tbl_style.add('BACKGROUND', (c, r), (c, r), colors.lightgreen)
                    elif val == 'FALTOU':
                        tbl_style.add('BACKGROUND', (c, r), (c, r), colors.salmon)
        t.setStyle(tbl_style)
        # footer callback
        def footer(canvas, doc):
            width, height = letter
            canvas.setFont("Helvetica", 8)
            canvas.drawString(40, 20, "Primeira Igreja Batista de Marília")
            canvas.drawCentredString(width/2, 20, "Face a Face de Homens")
            canvas.drawRightString(width-40, 20, str(doc.page))
        doc.build([t], onFirstPage=footer, onLaterPages=footer)
        bio.seek(0)
        return send_file(bio, download_name=f"{safe_event}.pdf",
                         as_attachment=True,
                         mimetype='application/pdf')
    else:
        abort(404)


@bp.route('/meetings/<int:meeting_id>/delete', methods=['POST'])
@login_required
def delete_meeting(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    # remove associated attendances and qrcodes manually
    from sqlalchemy import text
    db.session.execute(text('DELETE FROM attendance WHERE meeting_id = :mid'), {'mid': meeting_id})
    db.session.execute(text('DELETE FROM qr_code WHERE meeting_id = :mid'), {'mid': meeting_id})
    event_id = mt.event_id
    db.session.delete(mt)
    db.session.commit()
    flash('Reunião excluída', 'success')
    return redirect(url_for('admin.event_detail', event_id=event_id))


@bp.route('/meetings/<int:meeting_id>/qrcode', methods=['POST'])
@login_required
def generate_qrcode(meeting_id):
    mt = Meeting.query.get_or_404(meeting_id)
    # deactivate any existing active codes for this meeting
    QRCode.query.filter_by(meeting_id=mt.id, active=True).update({'active': False})
    # generate unique token
    token = secrets.token_urlsafe(16)
    q = QRCode(meeting=mt, token=token, active=True)
    db.session.add(q)
    db.session.commit()
    flash('QR code gerado (anterior redirecionará para este)', 'success')
    return redirect(url_for('admin.meeting_detail', meeting_id=meeting_id))


@bp.route('/qrcode/image/<int:qrcode_id>')
@login_required
def qrcode_image(qrcode_id):
    qr = QRCode.query.get_or_404(qrcode_id)
    # build URL using configured server address to avoid localhost
    from flask import current_app
    base = current_app.config.get('SERVER_ADDRESS')
    url = f"{base}/scan/{qr.token}"
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    response = make_response(buf.read())
    response.headers.set('Content-Type', 'image/png')
    return response

@bp.route('/qrcode/view/<int:qrcode_id>')
@login_required
def qrcode_view(qrcode_id):
    qr = QRCode.query.get_or_404(qrcode_id)
    event = qr.meeting.event
    return render_template('qrcode_view.html', qr=qr, event=event)


@bp.route('/qrcode/print/<int:qrcode_id>')
@login_required
def qrcode_print(qrcode_id):
    qr = QRCode.query.get_or_404(qrcode_id)
    event = qr.meeting.event
    return render_template('qrcode_print.html', qr=qr, event=event)

@bp.route('/qrcode/toggle/<int:qrcode_id>', methods=['POST'])
@login_required
def qrcode_toggle(qrcode_id):
    qr = QRCode.query.get_or_404(qrcode_id)
    qr.active = not qr.active
    db.session.commit()
    flash('QR code atualizado', 'success')
    return redirect(url_for('admin.meeting_detail', meeting_id=qr.meeting.id))
